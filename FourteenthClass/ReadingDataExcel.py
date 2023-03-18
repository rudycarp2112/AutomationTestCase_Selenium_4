# import openpyexcel
import openpyxl

# Structure of excel file to access it
# File -> Workbook -> Sheets -> Rows -> Cells

file="C:\data.xltx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Hoja1"]

rows=sheet.max_row
cols=sheet.max_column

#reading all columns for each row of excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="   ")
    print("")