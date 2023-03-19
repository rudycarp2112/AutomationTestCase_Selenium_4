import openpyxl
# https://youtu.be/S0WEaFucijs?list=PLUDwpEzHYYLsuUBvuoYTlN0KsBB5t-BDa&t=1893

# file="C:\\Users\\RudyX\\Desktop\\Writing.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Hoja1"]
#
# # Writing the same data
# for r in range(2,7):
#     for c in range(1,6):
#         sheet.cell(r,c).value="Franco"
#
# workbook.save(file)

# Writing multiple data

file="C:\\Users\\RudyX\\Desktop\\Writing2.xltx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Hoja1"]

sheet.cell(1,1).value=1
sheet.cell(1,2).value="Smith"
sheet.cell(1,2).value="Engineer"

sheet.cell(2,1).value=2
sheet.cell(2,2).value="Franco"
sheet.cell(2,3).value="Architect"

sheet.cell(3,1).value=3
sheet.cell(3,2).value="Rudy"
sheet.cell(3,3).value="Developer"

workbook.save(file) # save the file after entry the data

rows=sheet.max_row
cols=sheet.max_column

#reading all columns for each row of excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="   ")
    print("")